from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime, timezone
from app.database import get_db
from app.models.models import (
    Group,
    Subject,
    Teacher,
    Room,
    TimeSlot,
    LessonType,
    User,
    UserRole,
    ScheduleRow,
    ScheduleGenerationRun,
)
from app.schemas.schemas import (
    ScheduleGenerateResponse,
    ScheduleVersionResponse,
    ScheduleVersionUpdate,
    ScheduleGenerateRequest,
)
from app.dependencies import require_admin_or_dispatcher, require_authenticated
from app.routers.audit import log_action
from app.services.scheduler import ScheduleGenerator
from fastapi.responses import StreamingResponse

router = APIRouter(prefix="/api/schedule", tags=["Расписание"])

@router.get("/my", response_model=List[dict])
def get_my_schedule(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_authenticated),
):
    # [FIX F-01/U-02] Return full schedule details, not just IDs
    if current_user.role == UserRole.STUDENT:
        if not current_user.group_id:
            raise HTTPException(status_code=400, detail="Студент не привязан к группе")
        rows = db.query(ScheduleRow).filter(ScheduleRow.group_id == current_user.group_id).all()
    elif current_user.role == UserRole.TEACHER:
        if not current_user.teacher_id:
            raise HTTPException(status_code=400, detail="Не привязан к преподавателю")
        rows = db.query(ScheduleRow).filter(ScheduleRow.teacher_id == current_user.teacher_id).all()
    else:
        raise HTTPException(status_code=403, detail="Недоступно для этой роли")

    if not rows:
        return []

    # Bulk-load lookup tables to avoid N+1
    subject_map = {s.subject_id: s for s in db.query(Subject).all()}
    teacher_map = {t.teacher_id: t for t in db.query(Teacher).all()}
    room_map = {r.room_id: r for r in db.query(Room).all()}
    lt_map = {lt.lesson_type_id: lt for lt in db.query(LessonType).all()}
    ts_map = {ts.time_slot_id: ts for ts in db.query(TimeSlot).all()}
    group_map = {g.group_id: g for g in db.query(Group).all()}

    result = []
    for r in rows:
        ts = ts_map.get(int(r.time_slot_id))
        subj = subject_map.get(int(r.subject_id))
        tch = teacher_map.get(int(r.teacher_id))
        room = room_map.get(int(r.room_id))
        lt = lt_map.get(int(r.lesson_type_id))
        grp = group_map.get(int(r.group_id))
        result.append({
            "id": int(r.schedule_id),
            "group_name": grp.name if grp else "—",
            "subject_name": subj.name if subj else "—",
            "teacher_name": tch.full_name if tch else "—",
            "classroom_name": room.code if room else "—",
            "day_of_week": int(ts.day_of_week) if ts else 0,
            "time_slot_number": int(ts.slot_number) if ts else 0,
            "start_time": ts.start_time.strftime("%H:%M") if ts else "00:00",
            "end_time": ts.end_time.strftime("%H:%M") if ts else "00:00",
            "lesson_type": lt.name if lt else "—",
        })
    return result


# ──── Версии расписания ────

@router.get("/versions", response_model=List[ScheduleVersionResponse])
def get_versions(semester_id: Optional[int] = None, db: Session = Depends(get_db)):
    query = db.query(ScheduleGenerationRun)
    if semester_id:
        query = query.filter(ScheduleGenerationRun.academic_period_id == semester_id)
    runs = query.order_by(ScheduleGenerationRun.requested_at.desc()).all()
    return [
        {
            "id": int(r.generation_run_id),
            "semester_id": int(r.academic_period_id),
            "status": (r.parameters or {}).get("ui_status") or "generated",
            "description": r.notes,
            "created_at": r.requested_at,
        }
        for r in runs
    ]


@router.get("/versions/{version_id}", response_model=ScheduleVersionResponse)
def get_version(version_id: int, db: Session = Depends(get_db)):
    run = db.query(ScheduleGenerationRun).filter(ScheduleGenerationRun.generation_run_id == version_id).first()
    if not run:
        raise HTTPException(status_code=404, detail="Версия не найдена")
    return {
        "id": int(run.generation_run_id),
        "semester_id": int(run.academic_period_id),
        "status": (run.parameters or {}).get("ui_status") or "generated",
        "description": run.notes,
        "created_at": run.requested_at,
    }


@router.put("/versions/{version_id}", response_model=ScheduleVersionResponse)
def update_version(
    version_id: int, 
    data: ScheduleVersionUpdate, 
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_dispatcher)
):
    run = db.query(ScheduleGenerationRun).filter(ScheduleGenerationRun.generation_run_id == version_id).first()
    if not run:
        raise HTTPException(status_code=404, detail="Версия не найдена")
    
    payload = data.model_dump(exclude_unset=True)
    if "status" in payload and payload["status"] is not None:
        params = dict(run.parameters or {})
        params["ui_status"] = payload["status"]
        run.parameters = params

        if payload["status"] == "published":
            # Снимаем публикацию с других версий этого семестра (в UI-статусе)
            others = db.query(ScheduleGenerationRun).filter(
                ScheduleGenerationRun.academic_period_id == run.academic_period_id,
                ScheduleGenerationRun.generation_run_id != version_id
            ).all()
            for o in others:
                p = dict(o.parameters or {})
                p["ui_status"] = "archived"
                o.parameters = p
    if "description" in payload and payload["description"] is not None:
        run.notes = payload["description"]
    
    db.commit()
    db.refresh(run)
    return {
        "id": int(run.generation_run_id),
        "semester_id": int(run.academic_period_id),
        "status": (run.parameters or {}).get("ui_status") or "generated",
        "description": run.notes,
        "created_at": run.requested_at,
    }


@router.delete("/versions/{version_id}", status_code=204)
def delete_version(
    version_id: int, 
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_dispatcher)
):
    run = db.query(ScheduleGenerationRun).filter(ScheduleGenerationRun.generation_run_id == version_id).first()
    if not run:
        raise HTTPException(status_code=404, detail="Версия не найдена")
    # [FIX B-08] Save data before delete
    audit_info = {"description": run.notes}
    db.query(ScheduleRow).filter(ScheduleRow.source_run_id == version_id).delete()
    db.delete(run)
    db.commit()
    log_action(db, current_user.id, "DELETE", "schedule_generation_runs", version_id, audit_info)


@router.get("/versions/{version_id}/entries/detailed", response_model=List[dict])
def get_version_entries_detailed(
    version_id: int,
    group_id: Optional[int] = None,
    subject_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    query = db.query(ScheduleRow).filter(ScheduleRow.source_run_id == version_id)
    if group_id:
        query = query.filter(ScheduleRow.group_id == group_id)
    if subject_id:
        query = query.filter(ScheduleRow.subject_id == subject_id)

    rows = query.order_by(ScheduleRow.time_slot_id).all()

    group_map = {g.group_id: g for g in db.query(Group).all()}
    subject_map = {s.subject_id: s for s in db.query(Subject).all()}
    teacher_map = {t.teacher_id: t for t in db.query(Teacher).all()}
    room_map = {r.room_id: r for r in db.query(Room).all()}
    lt_map = {lt.lesson_type_id: lt for lt in db.query(LessonType).all()}
    ts_map = {ts.time_slot_id: ts for ts in db.query(TimeSlot).all()}

    result: list[dict] = []
    for r in rows:
        ts = ts_map.get(int(r.time_slot_id))
        subj = subject_map.get(int(r.subject_id))
        tch = teacher_map.get(int(r.teacher_id))
        room = room_map.get(int(r.room_id))
        lt = lt_map.get(int(r.lesson_type_id))
        grp = group_map.get(int(r.group_id))
        result.append(
            {
                "id": int(r.schedule_id),
                "group_id": int(r.group_id),
                "group_name": grp.name if grp else "—",
                "day_of_week": int(ts.day_of_week) if ts else 0,
                "time_slot_number": int(ts.slot_number) if ts else 0,
                "start_time": ts.start_time.strftime("%H:%M") if ts else "00:00",
                "end_time": ts.end_time.strftime("%H:%M") if ts else "00:00",
                "subject_name": subj.name if subj else "—",
                "teacher_id": int(r.teacher_id) if r.teacher_id else None,
                "teacher_name": tch.full_name if tch else "—",
                "classroom_id": int(r.room_id) if r.room_id else None,
                "classroom_name": room.code if room else "—",
                "week_type": "обе",
                "lesson_type": lt.name if lt else "—",
                "is_locked": False,
            }
        )
    return result


@router.put("/entries/{entry_id}")
def update_schedule_entry(
    entry_id: int,
    data: dict,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_dispatcher)
):
    from sqlalchemy import text as sa_text

    row = db.query(ScheduleRow).filter(ScheduleRow.schedule_id == entry_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Запись не найдена")

    new_teacher_id = data.get("teacher_id") or row.teacher_id
    new_room_id    = data.get("room_id")    or row.room_id
    new_ts_id      = row.time_slot_id

    if "day_of_week" in data and "slot_number" in data:
        ts = db.query(TimeSlot).filter(
            TimeSlot.day_of_week == data["day_of_week"],
            TimeSlot.slot_number == data["slot_number"]
        ).first()
        if not ts:
            raise HTTPException(status_code=400, detail="Временной слот не найден")
        new_ts_id = ts.time_slot_id

    # Pre-check conflicts to give user-friendly errors before hitting DB constraints
    if new_ts_id != row.time_slot_id or new_teacher_id != row.teacher_id or new_room_id != row.room_id:
        # Check if same group already has a lesson at the target slot (excluding this entry)
        group_conflict = db.query(ScheduleRow).filter(
            ScheduleRow.group_id == row.group_id,
            ScheduleRow.time_slot_id == new_ts_id,
            ScheduleRow.academic_period_id == row.academic_period_id,
            ScheduleRow.schedule_id != entry_id,
        ).first()
        if group_conflict:
            subj = db.query(Subject).filter(Subject.subject_id == group_conflict.subject_id).first()
            ts_obj = db.query(TimeSlot).filter(TimeSlot.time_slot_id == new_ts_id).first()
            day_names = {1: "Пн", 2: "Вт", 3: "Ср", 4: "Чт", 5: "Пт", 6: "Сб"}
            slot_label = f"{day_names.get(ts_obj.day_of_week,'?')} {ts_obj.slot_number}-я пара" if ts_obj else "этот слот"
            raise HTTPException(
                status_code=409,
                detail=f"У группы уже есть занятие в {slot_label}: «{subj.name if subj else '—'}». Выберите другое время."
            )

        # Check teacher conflict
        if new_teacher_id:
            teacher_conflict = db.query(ScheduleRow).filter(
                ScheduleRow.teacher_id == new_teacher_id,
                ScheduleRow.time_slot_id == new_ts_id,
                ScheduleRow.academic_period_id == row.academic_period_id,
                ScheduleRow.schedule_id != entry_id,
            ).first()
            if teacher_conflict:
                tch = db.query(Teacher).filter(Teacher.teacher_id == new_teacher_id).first()
                raise HTTPException(
                    status_code=409,
                    detail=f"Преподаватель «{tch.full_name if tch else '—'}» уже занят в это время."
                )

        # Check room conflict
        if new_room_id:
            room_conflict = db.query(ScheduleRow).filter(
                ScheduleRow.room_id == new_room_id,
                ScheduleRow.time_slot_id == new_ts_id,
                ScheduleRow.academic_period_id == row.academic_period_id,
                ScheduleRow.schedule_id != entry_id,
            ).first()
            if room_conflict:
                rm = db.query(Room).filter(Room.room_id == new_room_id).first()
                raise HTTPException(
                    status_code=409,
                    detail=f"Аудитория «{rm.code if rm else '—'}» уже занята в это время."
                )

    try:
        db.execute(sa_text("SET LOCAL session_replication_role = 'replica'"))
    except Exception:
        db.rollback()

    try:
        db.execute(
            sa_text(
                "UPDATE schedule SET teacher_id = :tid, room_id = :rid, "
                "time_slot_id = :tsid, updated_at = NOW() "
                "WHERE id = :eid"
            ),
            {"tid": new_teacher_id, "rid": new_room_id, "tsid": new_ts_id, "eid": entry_id}
        )
        db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail=f"Не удалось сохранить: {exc}")

    log_action(db, current_user.id, "UPDATE", "schedule", entry_id, {"desc": "Manual edit entry"})
    return {"status": "ok"}


@router.delete("/entries/{entry_id}", status_code=204)
def delete_schedule_entry(
    entry_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_dispatcher)
):
    row = db.query(ScheduleRow).filter(ScheduleRow.schedule_id == entry_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Запись не найдена")
    
    db.delete(row)
    db.commit()
    log_action(db, current_user.id, "DELETE", "schedule", entry_id, {"desc": "Delete entry"})
    return

# ──── Генерация ────

@router.post("/generate", response_model=ScheduleGenerateResponse)
def generate_schedule(
    data: ScheduleGenerateRequest,
    db: Session = Depends(get_db), 
    current_user: User = Depends(require_admin_or_dispatcher)
):
    try:
        run = ScheduleGenerationRun(
            academic_period_id=data.semester_id,
            status="queued",
            requested_by=current_user.username,
            generator_version="simple-v1",
            parameters={"description": data.description, "ui_status": "generated"},
            created_schedule_rows=0,
            detected_conflicts=0,
            requested_at=datetime.now(timezone.utc),
            started_at=datetime.now(timezone.utc),
            notes=data.description,
        )
        db.add(run)
        db.commit()
        db.refresh(run)

        # Удаляем строки расписания из предыдущих запусков для этого периода,
        # чтобы не нарушать unique-constraint (academic_period_id, room_id, time_slot_id).
        db.query(ScheduleRow).filter(
            ScheduleRow.academic_period_id == data.semester_id
        ).delete(synchronize_session=False)
        db.commit()

        generator = ScheduleGenerator(db)
        run.status = "running"
        db.commit()

        res = generator.generate(academic_period_id=data.semester_id, generation_run_id=run.generation_run_id)
        run.status = "completed"
        db.commit()
        log_action(db, current_user.id, "GENERATE", "schedule", int(run.generation_run_id),
                   {"semester_id": data.semester_id, "placed": res.placed_count, "total": res.total_count})

        return ScheduleGenerateResponse(
            version_id=int(run.generation_run_id),
            placed_count=res.placed_count,
            total_count=res.total_count,
            unplaced=res.unplaced,
            warnings=res.warnings
        )
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка генерации: {str(e)}")


# ──── Экспорт в формате оригинального расписания ────

@router.get("/versions/{version_id}/export")
def export_schedule_version(version_id: int, db: Session = Depends(get_db)):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    run = db.query(ScheduleGenerationRun).filter(ScheduleGenerationRun.generation_run_id == version_id).first()
    if not run:
        raise HTTPException(status_code=404, detail="Версия не найдена")

    rows = db.query(ScheduleRow).filter(ScheduleRow.source_run_id == version_id).all()

    subject_map = {s.subject_id: s for s in db.query(Subject).all()}
    teacher_map = {t.teacher_id: t for t in db.query(Teacher).all()}
    room_map = {r.room_id: r for r in db.query(Room).all()}
    lt_map = {lt.lesson_type_id: lt for lt in db.query(LessonType).all()}
    ts_map = {ts.time_slot_id: ts for ts in db.query(TimeSlot).all()}
    group_map = {g.group_id: g for g in db.query(Group).all()}

    # Collect unique groups ordered by name
    group_ids_in_schedule = sorted(
        set(int(r.group_id) for r in rows),
        key=lambda gid: group_map[gid].name if gid in group_map else ""
    )
    groups_list = [group_map[gid] for gid in group_ids_in_schedule if gid in group_map]

    # Build lookup: (day, slot, group_id) -> cell data dict
    sched: dict = {}
    for r in rows:
        ts = ts_map.get(int(r.time_slot_id))
        if not ts:
            continue
        key = (int(ts.day_of_week), int(ts.slot_number), int(r.group_id))
        subj = subject_map.get(int(r.subject_id))
        tch = teacher_map.get(int(r.teacher_id))
        room = room_map.get(int(r.room_id))
        lt = lt_map.get(int(r.lesson_type_id))
        sched[key] = {
            "subject": subj.name if subj else "—",
            "teacher": tch.full_name if tch else "—",
            "room": room.code if room else "—",
            "lt": lt.name if lt else "",
        }

    TIME_LABELS = {1: "08.00-09.30", 2: "09.35-11.05", 3: "11.20-12.50", 4: "13.10-14.40"}
    DAY_NAMES = {1: "Понедельник", 2: "Вторник", 3: "Среда", 4: "Четверг", 5: "Пятница", 6: "Суббота"}

    days_in_sched = sorted(set(int(ts_map[int(r.time_slot_id)].day_of_week)
                               for r in rows if int(r.time_slot_id) in ts_map)) or list(range(1, 7))
    slots_in_sched = sorted(set(int(ts_map[int(r.time_slot_id)].slot_number)
                                for r in rows if int(r.time_slot_id) in ts_map)) or [1, 2, 3, 4]

    # ── Styles ──────────────────────────────────────────────────────────────
    thin = Side(style="thin")
    medium = Side(style="medium")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    med_border = Border(left=medium, right=medium, top=medium, bottom=medium)

    fill_hdr      = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    fill_day      = PatternFill(start_color="2E4B7A", end_color="2E4B7A", fill_type="solid")
    fill_time     = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    fill_sat_day  = PatternFill(start_color="7B2020", end_color="7B2020", fill_type="solid")
    fill_sat_time = PatternFill(start_color="FDECEA", end_color="FDECEA", fill_type="solid")
    fill_white    = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_sat_cell = PatternFill(start_color="FFF5F5", end_color="FFF5F5", fill_type="solid")

    font_hdr      = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    font_title    = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    font_day      = Font(bold=True, color="FFFFFF", size=9,  name="Calibri")
    font_time     = Font(bold=True, color="1F3864", size=8,  name="Calibri")
    font_sat_time = Font(bold=True, color="7B2020", size=8,  name="Calibri")

    align_center   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left_top = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    align_vert     = Alignment(horizontal="center", vertical="center", text_rotation=90)

    wb = Workbook()
    GROUPS_PER_SHEET = 10
    sheet_count = max(1, -(-len(groups_list) // GROUPS_PER_SHEET))

    for sheet_idx in range(sheet_count):
        sheet_groups = groups_list[sheet_idx * GROUPS_PER_SHEET: (sheet_idx + 1) * GROUPS_PER_SHEET]
        if not sheet_groups:
            break

        ws = wb.active if sheet_idx == 0 else wb.create_sheet()
        ws.title = f"Расписание {sheet_idx + 1}"

        # Column widths
        ws.column_dimensions["A"].width = 14   # Day
        ws.column_dimensions["B"].width = 6    # Slot #
        ws.column_dimensions["C"].width = 13   # Time
        for i in range(len(sheet_groups)):
            ws.column_dimensions[get_column_letter(i + 4)].width = 27

        total_cols = 3 + len(sheet_groups)

        # Row 1: title
        ws.row_dimensions[1].height = 26
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        c = ws.cell(row=1, column=1,
                    value=f"Расписание занятий  ·  2 семестр 2025–2026  ·  Версия {version_id}")
        c.font = font_title; c.fill = fill_hdr; c.alignment = align_center; c.border = med_border

        # Row 2: headers
        ws.row_dimensions[2].height = 28
        for col, lbl in enumerate(["День", "№", "Расписание звонков"], start=1):
            c = ws.cell(row=2, column=col, value=lbl)
            c.font = font_hdr; c.fill = fill_hdr; c.alignment = align_center; c.border = thin_border
        for i, g in enumerate(sheet_groups):
            c = ws.cell(row=2, column=4 + i, value=g.name)
            c.font = font_hdr; c.fill = fill_hdr; c.alignment = align_center; c.border = thin_border

        # Data rows
        current_row = 3
        for day in days_in_sched:
            is_sat = day == 6
            day_start = current_row

            for si, slot in enumerate(slots_in_sched):
                ws.row_dimensions[current_row].height = 54

                # Day (write only on first slot; merge later)
                if si == 0:
                    dc = ws.cell(row=current_row, column=1, value=DAY_NAMES.get(day, str(day)))
                    dc.font = font_day
                    dc.fill = fill_sat_day if is_sat else fill_day
                    dc.border = thin_border

                # Slot number
                nc = ws.cell(row=current_row, column=2, value=slot)
                nc.font = Font(bold=True, size=10, name="Calibri",
                               color="C0331A" if is_sat else "1F3864")
                nc.fill = fill_sat_time if is_sat else fill_time
                nc.alignment = align_center; nc.border = thin_border

                # Time label
                tc = ws.cell(row=current_row, column=3, value=TIME_LABELS.get(slot, ""))
                tc.font = font_sat_time if is_sat else font_time
                tc.fill = fill_sat_time if is_sat else fill_time
                tc.alignment = align_center; tc.border = thin_border

                # Group cells
                for gi, g in enumerate(sheet_groups):
                    entry = sched.get((day, slot, g.group_id))
                    col = 4 + gi
                    cell = ws.cell(row=current_row, column=col)
                    cell.fill = fill_sat_cell if is_sat else fill_white
                    cell.border = thin_border
                    cell.alignment = align_left_top
                    if entry:
                        lines = [entry["subject"]]
                        parts = []
                        if entry["room"] and entry["room"] != "—":
                            parts.append(f"Каб. {entry['room']}")
                        if entry["teacher"] and entry["teacher"] != "—":
                            parts.append(entry["teacher"])
                        if parts:
                            lines.append("  |  ".join(parts))
                        if entry["lt"]:
                            lines.append(f"[{entry['lt']}]")
                        cell.value = "\n".join(lines)
                        cell.font = Font(size=8, name="Calibri",
                                         color="5C1A1A" if is_sat else "0D1B3E")

                current_row += 1

            # Merge day cell
            if len(slots_in_sched) > 1:
                ws.merge_cells(start_row=day_start, start_column=1,
                                end_row=current_row - 1, end_column=1)
            ws.cell(row=day_start, column=1).alignment = align_vert

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"schedule_v{version_id}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


